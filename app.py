import streamlit as st
import pandas as pd
import re
import io
import unicodedata
import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2 import service_account
import base64
from jinja2 import Template
import streamlit.components.v1 as components
import json
from datetime import datetime

# ========== CONFIG & SETUP ==========
FOLDER_ID_DEFAULT = "1AH34e-4R2gsNzX9q1lCBq8yoTIg3uCbr"
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

st.set_page_config(page_title="📋 Quản lý lớp học - VIAGS", layout="wide")
st.title("📋 Quản lý lớp học - VIAGS")

# ========== IMPORT JSON & EXPORT JSON ==========

if "hide_import_json" not in st.session_state:
    st.session_state["hide_import_json"] = False

col_import, col_export = st.columns([8, 2])
with col_import:
    if not st.session_state["hide_import_json"]:
        with st.expander("📂 Import dữ liệu đã lưu (JSON)", expanded=True):
            json_up = st.file_uploader("", type="json", key="import_json", label_visibility="collapsed")
            if json_up is not None:
                try:
                    raw = json.load(json_up)
                    new_data = {}
                    for name, content in raw.items():
                        ci = content.get("class_info", {})
                        df = pd.DataFrame(content.get("ds_hocvien", []))
                        for c in ["Mã NV", "Họ tên", "Đơn vị", "Điểm LT", "Điểm TH"]:
                            if c not in df.columns:
                                df[c] = ""
                        df = df[["Mã NV", "Họ tên", "Đơn vị", "Điểm LT", "Điểm TH"]]
                        new_data[name] = {
                            "class_info": ci,
                            "ds_hocvien": df
                        }
                    st.session_state["danh_sach_lop"] = new_data
                    current = st.session_state.get("ten_lop_hien_tai", "")
                    if current not in new_data:
                        keys = list(new_data.keys())
                        st.session_state["ten_lop_hien_tai"] = keys[-1] if keys else ""
                    msg = st.success("✅ Đã load JSON đầy đủ học viên và điểm.")
                    msg.empty()
                    st.session_state["hide_import_json"] = True
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Lỗi khi đọc file JSON: {e}")

# ========== EXPORT JSON ==========
with col_export:
    file_name = f"viags_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    export_dict = {}
    for cls_name, cls_data in st.session_state.get("danh_sach_lop", {}).items():
        df = cls_data["ds_hocvien"].copy()
        # Đảm bảo đủ 5 cột trước khi xuất
        for c in ["Mã NV", "Họ tên", "Đơn vị", "Điểm LT", "Điểm TH"]:
            if c not in df.columns:
                df[c] = ""
        export_dict[cls_name] = {
            "class_info": cls_data["class_info"],
            "ds_hocvien": df[["Mã NV", "Họ tên", "Đơn vị", "Điểm LT", "Điểm TH"]]
                            .to_dict(orient="records")
        }
    st.download_button(
        label="📥 Lưu dữ liệu JSON tất cả lớp",
        data=json.dumps(export_dict, ensure_ascii=False, indent=2),
        file_name=file_name,
        mime="application/json",
        use_container_width=True
    )

# ========== Google Drive API (Service Account) ==========
@st.cache_resource
def get_drive_service():
    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)

drive_service = get_drive_service()

def list_excel_files(folder_id):
    results = drive_service.files().list(
        q=f"'{folder_id}' in parents and trashed=false and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        fields="files(id, name)"
    ).execute()
    return [(f['name'], f['id']) for f in results.get('files', [])]

def download_excel_from_drive(file_id):
    req = drive_service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    fh.seek(0)
    return fh

# ========== SESSION STATE INIT ==========
for key, default in [
    ("danh_sach_lop", {}),
    ("ten_lop_hien_tai", ""),
    ("hien_nhap_excel", False)
]:
    if key not in st.session_state:
        st.session_state[key] = default
# ==== Chuẩn bị biến tạm cho cơ chế lưu khi chuyển tab ====
if "active_tab" not in st.session_state:
    st.session_state["active_tab"] = "1️⃣ Thông tin lớp học"

# ========== HELPER FUNCTIONS ==========
def chuan_hoa_thoi_gian(time_str):
    match = re.match(r"(\d{1,2})-(\d{1,2})/(\d{1,2}/\d{4})", str(time_str))
    if match:
        ngay1, ngay2, thangnam = match.groups()
        return f"{ngay1},{ngay2}/{thangnam}"
    return str(time_str).strip()

def remove_vietnamese_accents(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize('NFD', s)
    s = s.encode('ascii', 'ignore').decode('utf-8')
    s = s.replace(' ', '').lower()
    return s

def normalize_name(s):
    s = str(s) if s is not None else ""
    s = s.split('-')[0].strip()
    s = s.replace('Đ', 'D').replace('đ', 'd')
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', '', s)
    s = s.lower()
    return s

def strip_accents(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize('NFD', s)
    s = s.replace('Đ', 'D').replace('đ', 'd')
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

def round_score_str(score_str):
    # Làm tròn từng điểm trong chuỗi, ví dụ "8/9.5/7.5" -> "8/10/8"
    scores = []
    for s in str(score_str).replace(",", ".").split("/"):
        try:
            s_clean = re.sub(r"[^\d.]", "", s)
            if s_clean == "":
                continue
            f = float(s_clean)
            scores.append(str(int(f + 0.5)))
        except:
            pass
    return "/".join(scores)





def nhap_lop_tu_file(file_excel):
    wb = openpyxl.load_workbook(file_excel, data_only=True)
    so_lop_them = 0
    lop_moi_vua_them = None
    log_sheets = []
    for sheetname in wb.sheetnames:
        sheet_check = remove_vietnamese_accents(sheetname)
        if sheet_check == "mucluc":
            log_sheets.append(f"⏩ Bỏ qua sheet '{sheetname}' (Mục lục).")
            continue

        ws = wb[sheetname]
        ten_lop_goc = ws["D7"].value
        if not ten_lop_goc or str(ten_lop_goc).strip() == "":
            log_sheets.append(f"❌ Sheet '{sheetname}': Thiếu tên lớp ở D7.")
            continue

        thoi_gian = ws["D9"].value or ""
        thoi_gian_chuan = chuan_hoa_thoi_gian(thoi_gian)
        ten_lop = f"{str(ten_lop_goc).strip()}_{str(thoi_gian).strip()}"
        orig_ten_lop = ten_lop
        cnt = 1
        while ten_lop in st.session_state["danh_sach_lop"]:
            ten_lop = f"{orig_ten_lop}_{cnt}"
            cnt += 1

        loai_hinh_full = ws["B8"].value or ""
        if ":" in str(loai_hinh_full):
            loai_hinh = str(loai_hinh_full).split(":", 1)[-1].strip()
        else:
            loai_hinh = str(loai_hinh_full).strip()
        dia_diem = ws["D10"].value or ""

        data = []
        row = 14
        while True:
            ma_nv = ws[f"C{row}"].value
            ho_ten = ws[f"D{row}"].value
            don_vi = ws[f"E{row}"].value
            if (not ma_nv or str(ma_nv).strip() == "") and (not ho_ten or str(ho_ten).strip() == ""):
                break
            if any((
                (isinstance(ma_nv, str) and ("trưởng" in ma_nv.lower() or "trung tâm" in ma_nv.lower() or "ký tên" in ma_nv.lower())),
                (isinstance(ho_ten, str) and ("trưởng" in ho_ten.lower() or "trung tâm" in ho_ten.lower() or "ký tên" in ho_ten.lower())),
                (isinstance(don_vi, str) and ("trưởng" in don_vi.lower() or "trung tâm" in don_vi.lower() or "ký tên" in don_vi.lower()))
            )):
                break
            if (ma_nv and str(ma_nv).strip() != "") or (ho_ten and str(ho_ten).strip() != ""):
                data.append({
                    "Mã NV": str(ma_nv or "").strip(),
                    "Họ tên": str(ho_ten or "").strip(),
                    "Đơn vị": str(don_vi or "").strip(),
                    "Điểm": ""
                })
            row += 1

        if len(data) > 0:
            df = pd.DataFrame(data)
            st.session_state["danh_sach_lop"][ten_lop] = {
                "class_info": {
                    "course_name": ten_lop_goc,
                    "training_type": loai_hinh,
                    "time": thoi_gian_chuan,
                    "location": dia_diem,
                    "num_attended": "",
                    "num_total": "",
                },
                "ds_hocvien": df
            }
            lop_moi_vua_them = ten_lop
            so_lop_them += 1
            log_sheets.append(f"✅ Sheet '{sheetname}' ({ten_lop_goc}) đã nhập {len(data)} học viên (tên lớp: {ten_lop})")
        else:
            log_sheets.append(f"❌ Sheet '{sheetname}': Không có học viên ở C14-E14 trở đi.")

    if so_lop_them:
        st.session_state["ten_lop_hien_tai"] = lop_moi_vua_them
        cur_lop_data = st.session_state["danh_sach_lop"][lop_moi_vua_them]
        st.session_state["class_info_tmp"] = cur_lop_data.get("class_info", {}).copy()
        st.session_state["ds_hocvien_tmp"] = cur_lop_data.get("ds_hocvien", pd.DataFrame()).copy()
        st.session_state["diem_tmp"] = cur_lop_data.get("ds_hocvien", pd.DataFrame()).copy()
        st.success(f"Đã nhập xong {so_lop_them} lớp! Vào phần 'Chọn lớp' để kiểm tra.")
        for log in log_sheets:
            st.write(log)
        st.session_state["hien_nhap_excel"] = False
        st.rerun()
    else:
        for log in log_sheets:
            st.write(log)
        st.warning("Không sheet nào hợp lệ (phải có D7 là tên lớp và học viên từ C14-E14).")

def nhap_nhieu_lop_excel_modal():
    tab_drive, tab_file = st.tabs(["Từ Google Drive", "Từ máy tính"])
    # --- Tab 1: Google Drive ---
    with tab_drive:
        folder_id = FOLDER_ID_DEFAULT
        excel_files = list_excel_files(folder_id)
        file_map = {f[0]: f[1] for f in excel_files}
        if excel_files:
            selected_file = st.selectbox("Chọn file Excel danh sách lớp", list(file_map.keys()), key="select_drive_tab")
            if st.button("Tải và nhập từ Drive", key="btn_drive_import_tabdrive" + str(st.session_state.get("drive_tab_version", 0))):
                excel_bytes = download_excel_from_drive(file_map[selected_file])
                nhap_lop_tu_file(excel_bytes)
                # Reset key lần tiếp theo (chống đúp)
                st.session_state["drive_tab_version"] = st.session_state.get("drive_tab_version", 0) + 1
                st.session_state["hien_nhap_excel"] = False
                st.rerun()
        else:
            st.info("Không có file Excel nào trong folder Drive này.")
    # --- Tab 2: Máy tính ---
    with tab_file:
        file_excel = st.file_uploader(
            "Chọn file Excel danh sách lớp",
            type=["xlsx"],
            key="multi_class_uploader_import_tabfile" + str(st.session_state.get("file_tabfile_version", 0))
        )
        col_excel = st.columns([2, 1])
        with col_excel[0]:
            nhap_excel = st.button("Nhập các lớp vào hệ thống", key="btn_nhap_excel_tabfile" + str(st.session_state.get("file_tabfile_version", 0)))
        with col_excel[1]:
            huy_excel = st.button("❌ Đóng nhập nhiều lớp", key="btn_huy_excel_tabfile" + str(st.session_state.get("file_tabfile_version", 0)))
        if huy_excel:
            st.session_state["hien_nhap_excel"] = False
            # Reset key lần tiếp theo (chống đúp)
            st.session_state["file_tabfile_version"] = st.session_state.get("file_tabfile_version", 0) + 1
            st.rerun()
        if file_excel is not None and nhap_excel:
            nhap_lop_tu_file(file_excel)
            st.session_state["file_tabfile_version"] = st.session_state.get("file_tabfile_version", 0) + 1
            st.session_state["hien_nhap_excel"] = False
            st.rerun()
   
# ========== UI: Expander nhập nhiều lớp ==========
with st.expander("📥 Nhập nhiều lớp từ file Excel (mỗi sheet 1 lớp)", expanded=False):
    nhap_nhieu_lop_excel_modal()

# ========== UI: Quản lý nhiều lớp ==========
ds_lop = sorted(list(st.session_state["danh_sach_lop"].keys()), key=strip_accents)
chuc_nang = st.columns([5, 2, 1, 1])
with chuc_nang[0]:
    ten_lop = st.selectbox(
        "🗂️ Chọn lớp",
        ds_lop + ["+ Tạo lớp mới"],
        index=ds_lop.index(st.session_state["ten_lop_hien_tai"]) if st.session_state["ten_lop_hien_tai"] in ds_lop else len(ds_lop),
    )
with chuc_nang[1]:
    ten_moi = st.text_input("Tên lớp mới", value="", placeholder="VD: ATHK 01/2025")
with chuc_nang[2]:
    tao_lop = st.button("➕ Tạo lớp mới", use_container_width=True)
with chuc_nang[3]:
    if ds_lop and st.button("🗑️ Xóa lớp đang chọn", use_container_width=True):
        if st.session_state["ten_lop_hien_tai"] in st.session_state["danh_sach_lop"]:
            del st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]
            st.session_state["ten_lop_hien_tai"] = ds_lop[0] if ds_lop else ""
        st.success("Lớp đã được xóa thành công!")

if "ds_hocvien_tmp" not in st.session_state:
    if st.session_state["ten_lop_hien_tai"] and st.session_state["ten_lop_hien_tai"] in st.session_state["danh_sach_lop"]:
        st.session_state["ds_hocvien_tmp"] = st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]["ds_hocvien"].copy()
    else:
        st.session_state["ds_hocvien_tmp"] = pd.DataFrame({
            "Mã NV": [""] * 30,
            "Họ tên": [""] * 30,
            "Đơn vị": [""] * 30,
            "Điểm": [""] * 30
        })

if "diem_tmp" not in st.session_state:
    if st.session_state["ten_lop_hien_tai"] and st.session_state["ten_lop_hien_tai"] in st.session_state["danh_sach_lop"]:
        st.session_state["diem_tmp"] = st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]["ds_hocvien"].copy()
    else:
        st.session_state["diem_tmp"] = pd.DataFrame({
            "Mã NV": [""] * 30,
            "Họ tên": [""] * 30,
            "Đơn vị": [""] * 30,
            "Điểm": [""] * 30
        })

def save_data_when_switch_tab(new_tab):
    
    # Tab 1: Lưu thông tin lớp học khi chuyển tab
    if st.session_state["active_tab"] == "1️⃣ Thông tin lớp học" and new_tab != "1️⃣ Thông tin lớp học":
        st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]["class_info"] = st.session_state["class_info_tmp"].copy()
    # Tab 2: Lưu danh sách học viên (không điểm)
    if st.session_state["active_tab"] == "2️⃣ Danh sách học viên" and new_tab != "2️⃣ Danh sách học viên":
        ds = st.session_state["ds_hocvien_tmp"].copy()
        # Nếu đang có cột điểm thì reset, nếu không thì thôi (hoặc giữ lại, tùy bố muốn)
        for col in ["Điểm LT", "Điểm TH"]:
            if col in ds.columns:
                ds = ds.drop(columns=[col])
        # Reset lại điểm khi danh sách thay đổi
        ds["Điểm LT"] = ""
        ds["Điểm TH"] = ""
        st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]["ds_hocvien"] = ds.copy()
    # Tab 3: Lưu điểm
    if st.session_state["active_tab"] == "3️⃣ Cập nhật điểm" and new_tab != "3️⃣ Cập nhật điểm":
        st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]["ds_hocvien"] = st.session_state["diem_tmp"].copy()
    st.session_state["active_tab"] = new_tab

# Xử lý tạo lớp mới hoặc đổi lớp
if tao_lop and ten_moi.strip():
    if ten_moi not in st.session_state["danh_sach_lop"]:
        # Lưu dữ liệu lớp hiện tại trước khi chuyển sang lớp mới
        save_data_when_switch_tab(st.session_state["active_tab"])
        st.session_state["danh_sach_lop"][ten_moi] = {
            "class_info": {
                "course_name": "",
                "training_type": "",
                "time": "",
                "location": "",
                "num_attended": "",
                "num_total": "",
            },
            "ds_hocvien": pd.DataFrame({
                "Mã NV": [""] * 30,
                "Họ tên": [""] * 30,
                "Đơn vị": [""] * 30,
                "Điểm": [""] * 30
            }),
        }
        st.session_state["ten_lop_hien_tai"] = ten_moi
        cur_lop_data = st.session_state["danh_sach_lop"][ten_moi]
        st.session_state["class_info_tmp"] = cur_lop_data.get("class_info", {}).copy()
        st.session_state["ds_hocvien_tmp"] = cur_lop_data.get("ds_hocvien", pd.DataFrame()).copy()
        st.session_state["diem_tmp"] = cur_lop_data.get("ds_hocvien", pd.DataFrame()).copy()
        st.rerun()
    else:
        st.warning("Tên lớp đã tồn tại!")
elif ten_lop and ten_lop != "+ Tạo lớp mới":
    # Lưu dữ liệu lớp hiện tại trước khi chuyển sang lớp khác
    # Nếu đang ở tab 3 thì lưu điểm thủ công vào lớp cũ
    if st.session_state["active_tab"] == "3️⃣ Cập nhật điểm":
        st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]["ds_hocvien"] = st.session_state["diem_tmp"].copy()
    save_data_when_switch_tab(st.session_state["active_tab"])
    st.session_state["ten_lop_hien_tai"] = ten_lop
    cur_lop_data = st.session_state["danh_sach_lop"][ten_lop]
    st.session_state["class_info_tmp"] = cur_lop_data.get("class_info", {}).copy()
    st.session_state["ds_hocvien_tmp"] = cur_lop_data.get("ds_hocvien", pd.DataFrame()).copy()
    st.session_state["diem_tmp"] = cur_lop_data.get("ds_hocvien", pd.DataFrame()).copy()
# Nếu chưa có lớp nào, yêu cầu tạo trước
if not st.session_state["ten_lop_hien_tai"]:
    st.info("🔔 Hãy tạo lớp mới để bắt đầu nhập liệu và quản lý!")
    st.stop()

# Lấy dữ liệu lớp hiện tại
lop_data = st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]
class_info = lop_data.get("class_info", {})
ds_hocvien = lop_data.get("ds_hocvien", pd.DataFrame({
    "Mã NV": [""] * 30,
    "Họ tên": [""] * 30,
    "Đơn vị": [""] * 30,
    "Điểm": [""] * 30
}))




  
    
# ========== UI TABS ==========
tab1, tab2, tab3, tab4 = st.tabs([
    "1️⃣ Thông tin lớp học", 
    "2️⃣ Danh sách học viên",
    "3️⃣ Cập nhật điểm",
    "4️⃣ Chữ ký & xuất báo cáo"
])
# ========== Tab nội dung ==========

    
with tab1:
    save_data_when_switch_tab("1️⃣ Thông tin lớp học")
    st.subheader("📝 Nhập thông tin lớp học")
    ten_lop = st.session_state["ten_lop_hien_tai"]

    # Lấy dữ liệu cũ (nếu có), hoặc dùng mẫu
    cur_info = st.session_state["danh_sach_lop"][ten_lop].get("class_info", {})
    sample = [
        "An toàn hàng không",
        "Định kỳ/Elearning+Trực tiếp",
        "02/01/2025",
        "TTĐT MB",
        "VNBA25-ĐKVH04"
    ]
    default_value = "\n".join([
        cur_info.get("course_name", ""),
        cur_info.get("training_type", ""),
        cur_info.get("time", ""),
        cur_info.get("location", ""),
        cur_info.get("class_code", "")
    ]) if any(cur_info.values()) else "\n".join(sample)

    # Text area nhập 5 dòng
    txt = st.text_area(
        "Dán 5 dòng: Môn học, Loại hình, Thời gian, Địa điểm, Mã lớp (ghi chú)",
        value=default_value,
        height=130
    )

    # Parse từng dòng
    lines = txt.split("\n")
    new_info = {
        "course_name": lines[0].strip() if len(lines) > 0 else "",
        "training_type": lines[1].strip() if len(lines) > 1 else "",
        "time": lines[2].strip() if len(lines) > 2 else "",
        "location": lines[3].strip() if len(lines) > 3 else "",
        "class_code": lines[4].strip() if len(lines) > 4 else "",
    }

    # Lưu ngay vào session_state chính
    st.session_state["danh_sach_lop"][ten_lop]["class_info"] = new_info
    st.session_state["class_info_tmp"] = new_info.copy()
    st.info("✅ Đã lưu thông tin lớp học.")

with tab2:
    save_data_when_switch_tab("2️⃣ Danh sách học viên")
    st.subheader("📋 Danh sách học viên")
    ten_lop = st.session_state["ten_lop_hien_tai"]
    # Lấy toàn bộ DataFrame, giữ cả cột điểm nếu có
    df_all = st.session_state["danh_sach_lop"][ten_lop]["ds_hocvien"].copy()

    # Đảm bảo đủ 3 cột cơ bản
    for col in ["Mã NV", "Họ tên", "Đơn vị"]:
        if col not in df_all.columns:
            df_all[col] = ""

    # Chỉ hiển thị 3 cột để edit, nhưng không xóa các cột khác
    df_info = df_all[["Mã NV", "Họ tên", "Đơn vị"]]

    edited_info = st.data_editor(
        df_info,
        key=f"editor_ds_{ten_lop}",
        num_rows="dynamic",
        use_container_width=True,
        column_order=["Mã NV", "Họ tên", "Đơn vị"],
        column_config={
            "Mã NV": st.column_config.TextColumn(width="x-small"),
            "Họ tên": st.column_config.TextColumn(width="large"),
            "Đơn vị": st.column_config.TextColumn(width="medium"),
        }
    )

    # Cập nhật trở lại DataFrame chính, giữ nguyên các cột khác (ví dụ Điểm LT, Điểm TH)
    df_all.loc[:, ["Mã NV", "Họ tên", "Đơn vị"]] = edited_info

    # Lưu vào session_state
    st.session_state["danh_sach_lop"][ten_lop]["ds_hocvien"] = df_all


with tab3:
    save_data_when_switch_tab("3️⃣ Cập nhật điểm")
    st.subheader("📊 Nhập điểm LT / TH")
    ten_lop = st.session_state["ten_lop_hien_tai"]
    df = st.session_state["danh_sach_lop"][ten_lop]["ds_hocvien"].copy()

    # Đảm bảo đủ 5 cột
    for c in ["Mã NV", "Họ tên", "Đơn vị", "Điểm LT", "Điểm TH"]:
        if c not in df.columns:
            df[c] = ""
    df = df[["Mã NV", "Họ tên", "Đơn vị", "Điểm LT", "Điểm TH"]]

    # ----- Nhập điểm tự động từ file LMS -----
    col_lms, col_dotthi = st.columns(2)
    with col_lms:
        st.markdown("**<span style='font-size:16px'>📥 File điểm LMS</span>**", unsafe_allow_html=True)
        uploaded_lms = st.file_uploader("", type=["xlsx"], key="lms_tab3", label_visibility="collapsed")
    with col_dotthi:
        st.markdown("**<span style='font-size:16px'>📥 File điểm Đợt thi</span>**", unsafe_allow_html=True)
        uploaded_dotthi = st.file_uploader("", type=["xlsx"], key="dotthi_tab3", label_visibility="collapsed")
    # --- a) LMS ---
    if uploaded_lms:
        df_diem = pd.read_excel(uploaded_lms)
        col_name_hoten  = df_diem.columns[3]
        col_name_lanthi = df_diem.columns[6]
        df_diem["HoTenChuan"] = df_diem[col_name_hoten].apply(normalize_name)
        def extract_lms(txt):
            if not isinstance(txt, str): return ""
            lst = re.findall(r"Lần \d+\s*:\s*(\d+(?:\.\d+)?)", txt)
            return "/".join(lst)
        df_diem["DiemDaXuLy"] = df_diem[col_name_lanthi].apply(extract_lms)

        # Merge vào toàn bộ danh sách
        map_lms = dict(zip(df_diem["HoTenChuan"], df_diem["DiemDaXuLy"]))
        df["HoTenChuan"] = df["Họ tên"].apply(normalize_name)
        count = 0
        for i, row in df.iterrows():
            k = row["HoTenChuan"]
            if k in map_lms and map_lms[k]:
                df.at[i, "Điểm LT"] = round_score_str(map_lms[k])
                count += 1
        df = df.drop(columns=["HoTenChuan"])
        st.success(f"✅ Đã ghép điểm LT từ LMS cho {count} học viên.")

    # --- b) Đợt thi ---
    if uploaded_dotthi:
        df_dot = pd.read_excel(uploaded_dotthi)
        c_hot    = df_dot.columns[2]
        c_d1     = df_dot.columns[4]
        c_dn     = df_dot.columns[6]
        df_dot["HoTenChuan"] = df_dot[c_hot].apply(normalize_name)
        def extract_dotthi(row):
            if pd.notnull(row[c_dn]) and str(row[c_dn]).strip():
                m = re.findall(r"Lần\s*\d+\s*:\s*(\d+(?:\.\d+)?)", str(row[c_dn]))
                return "/".join(m) if m else str(row[c_dn]).strip()
            if pd.notnull(row[c_d1]) and str(row[c_d1]).strip():
                return str(row[c_d1]).strip()
            return ""
        df_dot["DiemDaXuLy"] = df_dot.apply(extract_dotthi, axis=1)

        map_dot = dict(zip(df_dot["HoTenChuan"], df_dot["DiemDaXuLy"]))
        df["HoTenChuan"] = df["Họ tên"].apply(normalize_name)
        cnt2 = 0
        for i, row in df.iterrows():
            k = row["HoTenChuan"]
            if k in map_dot and map_dot[k]:
                df.at[i, "Điểm LT"] = round_score_str(map_dot[k])
                cnt2 += 1
        df = df.drop(columns=["HoTenChuan"])
        st.success(f"✅ Đã ghép điểm LT từ Đợt thi cho {cnt2} học viên.")

    # ----- Chỉnh sửa thủ công toàn bộ học viên -----
    st.markdown("**✏️ Chỉnh sửa thủ công**")
    edited = st.data_editor(
        df,
        key=f"editor_diem_{ten_lop}",
        num_rows="fixed",
        use_container_width=True,
        disabled=["Mã NV", "Họ tên", "Đơn vị"]
    )
    # Lưu lại điểm toàn bộ học viên
    #st.session_state["danh_sach_lop"][ten_lop]["ds_hocvien"] = edited
    st.session_state["diem_tmp"] = edited.copy()
    st.session_state["danh_sach_lop"][ten_lop]["ds_hocvien"] = edited.copy()

with tab4:
    save_data_when_switch_tab("4️⃣ Chữ ký & xuất báo cáo")
    st.subheader("Thông tin chữ ký báo cáo & Xuất báo cáo")

    # Lấy danh sách học viên từ đúng lớp đang chọn
    ds_hocvien = st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]["ds_hocvien"].copy()

    with st.expander("✍️ Thông tin chữ ký báo cáo", expanded=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            gv_huong_dan = st.text_input("Giáo viên hướng dẫn", value="Nguyễn Đức Nghĩa")
        with col2:
            truong_bo_mon = st.text_input("Trưởng bộ môn", value="Ngô Trung Thành")
        with col3:
            truong_tt = st.text_input("Trưởng TTĐT", value="Nguyễn Chí Kiên")

    # Lấy thông tin lớp
    lop_data = st.session_state["danh_sach_lop"][st.session_state["ten_lop_hien_tai"]]
    class_info = lop_data["class_info"]
    course_name = class_info.get("course_name", "")
    training_type = class_info.get("training_type", "")
    time = class_info.get("time", "")
    location = class_info.get("location", "")
    num_attended = class_info.get("num_attended", "")
    num_total = class_info.get("num_total", "")

    def extract_days(time_str):
        # Nếu là datetime, chuyển về string
        if hasattr(time_str, "strftime"):
            time_str = time_str.strftime("%d/%m/%Y")
        time_str = str(time_str)
        if not time_str:
            return []
        time_str = time_str.replace('S', '').replace('s', '')
        match = re.search(r'([\d,]+)/(\d{1,2})/(\d{4})', time_str)
        if match:
            days = [d.strip() for d in match.group(1).split(',')]
            month = match.group(2)
            return [f"{d}/{month}" for d in days if d]
        match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', time_str)
        if match:
            return [f"{match.group(1)}/{match.group(2)}"]
        return []

    def ma_hoa_ten_lop(course_name, time_str):
        #Mã hóa tên lớp thành tên ngắn gọn (ví dụ: YTCN_040625)
        from datetime import datetime
        words = re.findall(r'\w+', str(course_name))
        initials = ''.join([w[0].upper() for w in words])[:10]
        # Nếu là datetime, chuyển về string
        if hasattr(time_str, "strftime"):
            s = time_str.strftime("%d/%m/%Y")
        else:
            s = str(time_str)
        match = re.match(r'(\d{1,2})[, -](\d{1,2})/(\d{1,2})/(\d{4})', s)
        if match:
            dd1 = match.group(1).zfill(2)
            dd2 = match.group(2).zfill(2)
            mm = match.group(3).zfill(2)
            yy = match.group(4)[-2:]
            time_part = f"{dd1}{dd2}{mm}{yy}"
        else:
            match2 = re.match(r'(\d{1,2})[, -](\d{1,2})/(\d{1,2})$', s)
            if match2:
                dd1 = match2.group(1).zfill(2)
                dd2 = match2.group(2).zfill(2)
                mm = match2.group(3).zfill(2)
                yy = str(datetime.now().year)[-2:]
                time_part = f"{dd1}{dd2}{mm}{yy}"
            else:
                match3 = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
                if match3:
                    dd = match3.group(1).zfill(2)
                    mm = match3.group(2).zfill(2)
                    yy = match3.group(3)[-2:]
                    time_part = f"{dd}{mm}{yy}"
                else:
                    nums = re.findall(r'\d+', s)
                    if len(nums) >= 4:
                        dd1 = nums[0].zfill(2)
                        dd2 = nums[2].zfill(2)
                        mm = nums[-2].zfill(2)
                        yy = nums[-1][-2:]
                        time_part = f"{dd1}{dd2}{mm}{yy}"
                    elif len(nums) == 3:
                        dd = nums[0].zfill(2)
                        mm = nums[1].zfill(2)
                        yy = nums[2][-2:]
                        time_part = f"{dd}{mm}{yy}"
                    else:
                        time_part = "000000"
        return f"{initials}_{time_part}"

    with open("logo_viags.png", "rb") as image_file:
        logo_base64 = base64.b64encode(image_file.read()).decode()

    col1, col2, _ = st.columns([1, 1, 4])
    with col1:
        bckq = st.button("📄In báo cáo kết quả")
    with col2:
        diem_danh = st.button("Tạo bảng điểm danh")

    if bckq:
        if ds_hocvien.empty:
            st.warning("Vui lòng nhập danh sách học viên!")
        else:
            ds_hocvien_filtered = ds_hocvien[
                (ds_hocvien["Mã NV"].astype(str).str.strip() != "") | (ds_hocvien["Họ tên"].astype(str).str.strip() != "")]
            data = []

            diem_lt_nonempty = ds_hocvien_filtered["Điểm LT"].astype(str).str.strip().replace("-", "").replace("nan", "").replace("None", "").ne("").sum()
            diem_th_nonempty = ds_hocvien_filtered["Điểm TH"].astype(str).str.strip().replace("-", "").replace("nan", "").replace("None", "").ne("").sum()
            use_5b = diem_lt_nonempty > 0 and diem_th_nonempty > 0
            template_file = "report_template_5b.html" if use_5b else "report_template_5a.html"

            for i, row in ds_hocvien_filtered.iterrows():
                ma_nv = str(row.get("Mã NV", "") or "").strip()
                ho_ten = str(row.get("Họ tên", "") or "").strip()
                if (not ma_nv or ma_nv.lower() == "none") and (not ho_ten or ho_ten.lower() == "none"):
                    continue

                diem_lt = str(row.get("Điểm LT", "") or "").strip()
                diem_th = str(row.get("Điểm TH", "") or "").strip()
                # Làm tròn điểm LT/TH
                diem_lt = round_score_str(diem_lt)
                diem_th = round_score_str(diem_th)
                if use_5b:
                    diem_lt = diem_lt if diem_lt not in ["", "nan", "None", None] else "-"
                    diem_th = diem_th if diem_th not in ["", "nan", "None", None] else "-"
                 
                    def get_last_score(s):
                        if s in ["", "-", "nan", "None", None]:
                            return 0
                        parts = [p.strip() for p in str(s).replace(",", ".").split("/") if p.strip().replace(".", "", 1).isdigit()]
                        return float(parts[-1]) if parts else 0

                    scores_lt = [s for s in str(diem_lt).split("/") if s.strip().isdigit()]
                    scores_th = [s for s in str(diem_th).split("/") if s.strip().isdigit()]
                    num_tests = max(len(scores_lt), len(scores_th))

                    # Kiểm tra có học viên nào đủ cả LT và TH không
                    all_have_both = any(
                        len([x for x in str(r.get("Điểm LT", "")).split("/") if x.strip().isdigit()]) > 0 and
                        len([x for x in str(r.get("Điểm TH", "")).split("/") if x.strip().isdigit()]) > 0
                        for _, r in ds_hocvien_filtered.iterrows()
                    )

                    # Nếu các học viên khác có đủ cả LT và TH mà học viên này chỉ có 1 môn
                    if all_have_both and ((len(scores_lt) == 0 and len(scores_th) > 0) or (len(scores_th) == 0 and len(scores_lt) > 0)):
                        diem_tb = "-"
                        xep_loai = "-"
                        note = "Thiếu môn"
                    elif diem_lt == "-" and diem_th == "-":
                        diem_tb = "-"
                        xep_loai = "-"
                        note = "Vắng"
                    else:
                        d_lt = get_last_score(diem_lt)
                        d_th = get_last_score(diem_th)
                        try:
                            diem_tb = int((d_lt + 2 * d_th) / 3 + 0.5)
                        except:
                            diem_tb = 0
                        # Xếp loại: 
                            # Nếu bất kỳ điểm cuối LT hoặc TH < 80 thì luôn Không đạt
                            # >1 lần chỉ "Đạt" nếu >=80, không có "Xuất sắc"
                        if d_lt < 80 or d_th < 80:
                            xep_loai = "Không đạt"
                        elif num_tests > 1:
                            xep_loai = "Đạt" if diem_tb >= 80 else "Không đạt"
                        else:
                            if diem_tb >= 95:
                                xep_loai = "Xuất sắc"
                            elif diem_tb >= 80:
                                xep_loai = "Đạt"
                            else:
                                xep_loai = "Không đạt"
                        # Ghi chú
                        note = ""
                        main_scores = scores_lt if len(scores_lt) > 1 else scores_th
                        if len(main_scores) > 1:
                            note = f"Kiểm tra lần {'/'.join(str(i+1) for i in range(len(main_scores)))}"

                    data.append({
                        "id": ma_nv,
                        "name": ho_ten,
                        "unit": str(row.get("Đơn vị", "") or "").strip(),
                        "score_lt": diem_lt,
                        "score_th": diem_th,
                        "score_tb": diem_tb,
                        "rank": xep_loai,
                        "note": note,
                        "num_tests": num_tests
                    })
                else:
                    if diem_lt and diem_lt not in ["", "-", "nan", "None", None]:
                        diem_chinh = diem_lt
                    elif diem_th and diem_th not in ["", "-", "nan", "None", None]:
                        diem_chinh = diem_th
                    else:
                        diem_chinh = "-"
                    try:
                        parts = [p.strip() for p in str(diem_chinh).replace(",", ".").split("/") if p.strip().replace(".", "", 1).isdigit()]
                        diem_num = float(parts[-1]) if parts else 0
                    except:
                        diem_num = 0
                    scores = [s for s in str(diem_chinh).split("/") if s.strip().isdigit()]
                    num_tests = len(scores)
                    if diem_chinh in ["", "nan", "None", None]:
                        diem_chinh = "-"
                    if diem_chinh == "-":
                        xep_loai = "-"
                    elif num_tests > 1:
                        xep_loai = "Đạt" if diem_num >= 80 else "Không đạt"
                    else:
                        if diem_num >= 95:
                            xep_loai = "Xuất sắc"
                        elif diem_num >= 80:
                            xep_loai = "Đạt"
                        elif diem_num > 0:
                            xep_loai = "Không đạt"
                        else:
                            xep_loai = "-"
                    # Ghi chú
                    if diem_chinh == "-":
                        note = "Vắng"
                    else:
                        if num_tests > 1:
                            note = f"Kiểm tra lần {'/'.join(str(i+1) for i in range(num_tests))}"
                        else:
                            note = ""
                    data.append({
                        "id": ma_nv,
                        "name": ho_ten,
                        "unit": str(row.get("Đơn vị", "") or "").strip(),
                        "score": diem_chinh,
                        "rank": xep_loai,
                        "note": note,
                        "num_tests": num_tests
                    })

            # Sắp xếp
            if use_5b:
                def get_group_5b(student, all_have_both):
                    scores_lt = [x for x in str(student.get("score_lt", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]
                    scores_th = [x for x in str(student.get("score_th", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]
                    n_lt = len(scores_lt)
                    n_th = len(scores_th)
                    total = n_lt + n_th
                    rank = student.get("rank", "-")

                    # 1: Xuất sắc
                    if rank == "Xuất sắc":
                        return 1
                    # 2: Đạt (1 LT, 1 TH)
                    if rank == "Đạt" and n_lt == 1 and n_th == 1:
                        return 2
                    # 3: Đạt (tổng số lần thi LT+TH = 3)
                    if rank == "Đạt" and total == 3:
                        return 3
                    # 4: Đạt (2 LT, 2 TH)
                    if rank == "Đạt" and n_lt == 2 and n_th == 2:
                        return 4
                    # 5: Đạt (1 LT/3 TH hoặc 3 LT/1 TH)
                    if rank == "Đạt" and ((n_lt == 1 and n_th == 3) or (n_lt == 3 and n_th == 1)):
                        return 5
                    # 6: Đạt (tổng 5 lần thi)
                    if rank == "Đạt" and total == 5:
                        return 6
                    # 7: Đạt (tổng 6 lần thi)
                    if rank == "Đạt" and total == 6:
                        return 7
                    # 8: Không đạt
                    if rank == "Không đạt":
                        return 8
                    # 9: Thiếu môn
                    if rank == "-" and ((n_lt == 0 and n_th > 0) or (n_th == 0 and n_lt > 0)) and all_have_both:
                        student["note"] = "Thiếu môn"
                        return 9
                    # 10: Vắng
                    if (n_lt == 0 and n_th == 0):
                        student["note"] = "Vắng"
                        return 10
                    return 11

                all_have_both = any(
                    len([x for x in str(s.get("score_lt", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]) > 0 and
                    len([x for x in str(s.get("score_th", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]) > 0
                    for s in data
                )

                for student in data:
                    student["group"] = get_group_5b(student, all_have_both)
                    # Lấy list điểm từng lần thi (float, từ trái sang phải)
                    scores = []
                    if "score_lt" in student:
                        scores += [float(x) for x in str(student.get("score_lt", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]
                    if "score_th" in student:
                        scores += [float(x) for x in str(student.get("score_th", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]
                    student["score_list"] = scores
                    student["last_score"] = scores[-1] if scores else 0

                # Sắp xếp theo group, điểm trung bình cao hơn trước, cùng điểm thì so sánh từng lần thi từ cuối về trước, rồi đến tên
                data_sorted = sorted(
                    data,
                    key=lambda row: (
                        row["group"],
                        # Nếu là group 9 (Thiếu môn) thì số lần thi ít hơn xếp trước
                        row["num_tests"] if row["group"] == 9
                        # Nếu là group 8 (Không đạt) thì số lần thi ít hơn xếp trước
                        else row["num_tests"] if row["group"] == 8
                        # Các group khác thì theo điểm trung bình giảm dần
                        else -(float(row.get("score_tb", 0)) if str(row.get("score_tb", 0)).replace('.', '', 1).isdigit() else -1),
                        tuple([-x for x in row.get("score_list", [0])[::-1]]),
                        row["name"]
                    )
                )
            else:
                def calc_group_numtests_score1(student):
                    rank = student.get("rank", "-")
                    if rank == "Xuất sắc":
                        group = 1
                    elif rank == "Đạt":
                        group = 2
                    elif rank == "Không đạt":
                        group = 3
                    else:
                        group = 4  # Vắng hoặc "-"
                    num_tests = student.get("num_tests", 1)
                    if "score_tb" in student:
                        try:
                            score_1 = float(student.get("score_tb", 0))
                        except:
                            score_1 = 0
                    else:
                        try:
                            score_1 = float(student.get("score", 0))
                        except:
                            score_1 = 0
                    return group, num_tests, score_1

                for student in data:
                    group, num_tests, score_1 = calc_group_numtests_score1(student)
                    student["group"] = group
                    student["num_tests"] = num_tests
                    student["score_1"] = score_1
                    if "score_lt" in student:
                        scores = [float(x) for x in str(student.get("score_lt", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]
                    elif "score" in student:
                        scores = [float(x) for x in str(student.get("score", "")).replace(",", ".").split("/") if x.strip().replace(".", "", 1).isdigit()]
                    else:
                        scores = []
                    student["last_score"] = scores[-1] if scores else 0
                    student["score_list"] = scores
                data_sorted = sorted(
                    data,
                    key=lambda row: (
                        row["group"],
                        row["num_tests"],
                        tuple([-x for x in row.get("score_list", [0])[::-1]]),
                        row["name"]
                    )
                )
            # Chia thành 2 trang nếu cần
            num_total = len(data_sorted)
            max_rows_per_page = 21  # Số hàng tối đa trên mỗi trang
            if num_total > max_rows_per_page:
                students_trang1 = data_sorted[:max_rows_per_page]
                students_trang2 = data_sorted[max_rows_per_page:]
            else:
                students_trang1 = data_sorted
                students_trang2 = []

            # Tính lại số lượng nếu chưa có
            if not num_attended or not num_total:
                num_total = len(data_sorted)
                if use_5b:
                    num_attended = sum(1 for x in data_sorted if str(x.get("score_lt", "") or "").strip() not in ["", "-", "nan", "None"] or str(x.get("score_th", "") or "").strip() not in ["", "-", "nan", "None"])
                else:
                    num_attended = sum(1 for x in data_sorted if str(x.get("score", "") or "").strip() not in ["", "-", "nan", "None"])

            # Xử lý ngày
            days = extract_days(time)
            for i, student in enumerate(data_sorted):
                student["day1"] = days[0] if len(days) > 0 else ""
                student["day2"] = days[1] if len(days) > 1 else ""
                student["day3"] = days[2] if len(days) > 2 else ""

            with open(template_file, "r", encoding="utf-8") as f:
                template_str = f.read()
            template = Template(template_str)
            # Định dạng thời gian
            if hasattr(time, "strftime"):
                time = time.strftime("%d/%m/%Y")
            else:
                time = str(time)
            
            # Render template
            rendered = template.render(
                students=data_sorted,
                students_trang1=students_trang1,
                students_trang2=students_trang2,
                course_name=course_name,
                training_type=training_type,
                time=time,
                location=location,
                num_attended=num_attended,
                num_total=num_total,
                class_code=class_info.get("class_code", ""),
                gv_huong_dan=gv_huong_dan,
                truong_bo_mon=truong_bo_mon,
                truong_tt=truong_tt,
                logo_base64=logo_base64,
                max_rows_per_page=max_rows_per_page
            )
            # Tìm số lần thi lớn nhất
            max_tests = 1
            for s in data_sorted:
                if "score_lt" in s:
                    scores = [x for x in str(s.get("score_lt", "")) .split("/") if x.strip().isdigit()]
                else:
                    scores = [x for x in str(s.get("score", "")) .split("/") if x.strip().isdigit()]
                if len(scores) > max_tests:
                    max_tests = len(scores)

            # Thêm cột điểm từng lần thi
            for s in data_sorted:
                if "score_lt" in s:
                    scores = [x for x in str(s.get("score_lt", "")) .split("/") if x.strip().isdigit()]
                else:
                    scores = [x for x in str(s.get("score", "")) .split("/") if x.strip().isdigit()]
                for i in range(max_tests):
                    s[f"Điểm lần {i+1}"] = scores[i] if i < len(scores) else ""
            
            # Xuất Excel đúng tên mã hóa
            file_basename = ma_hoa_ten_lop(course_name, time)
            file_excel = f"{file_basename}.xlsx"
            df_baocao = pd.DataFrame(data_sorted)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # Ghi bảng học viên từ dòng 7 (index=6) để phía trên trống cho info lớp
                df_baocao.to_excel(writer, index=False, sheet_name="Báo cáo", startrow=6)
                workbook  = writer.book
                worksheet = writer.sheets["Báo cáo"]

                # Ghi thông tin lớp học vào các dòng đầu
                worksheet.write("A1", "Môn/Khóa học:")
                worksheet.write("B1", course_name)
                worksheet.write("A2", "Loại hình/hình thức đào tạo:")
                worksheet.write("B2", training_type)
                worksheet.write("A3", "Thời gian:")
                worksheet.write("B3", time)
                worksheet.write("A4", "Địa điểm:")
                worksheet.write("B4", location)
                worksheet.write("A5", "Số học viên tham dự/tổng số học viên:")
                worksheet.write("B5", f"{num_attended}/{num_total}")
                worksheet.write("A6", "Mã lớp/Ghi chú:")
                worksheet.write("B6", class_info.get("class_code", ""))
            output.seek(0)
            excel_b64 = base64.b64encode(output.read()).decode()

            html_report = f"""
            <div style="text-align:right; margin-bottom:12px;" class="no-print">
                <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_b64}"
                download="{file_excel}"
                style="display:inline-block; font-size:18px; padding:6px 18px; margin-right:16px; background:#f0f0f0; border-radius:4px; text-decoration:none; border:1px solid #ccc;">
                📥 Tải báo cáo Excel
                </a>
                <button onclick="window.print()" style="font-size:18px;padding:6px 18px;">🖨️ In báo cáo kết quả</button>
            </div>
            {rendered}
            """
            st.subheader("📄 Xem trước báo cáo")
            st.components.v1.html(html_report, height=1200, scrolling=True)

    # Nếu có nút điểm danh, tạo bảng điểm danh
    if diem_danh:
        df = ds_hocvien[(ds_hocvien["Mã NV"].astype(str).str.strip() != "") | (ds_hocvien["Họ tên"].astype(str).str.strip() != "")]
        df = df.reset_index(drop=True)
        if hasattr(time, "strftime"):
            time = time.strftime("%d/%m/%Y")
        else:
            time = str(time)
        days = extract_days(time)
        students = []
        for i, row in df.iterrows():
            diem_lt = str(row.get("Điểm LT", "") or "").strip()
            check = "X" if diem_lt and diem_lt not in ["", "-", "None"] else "V"
            students.append({
                "stt": i + 1,
                "id": str(row.get("Mã NV", "") or "").strip(),
                "name": str(row.get("Họ tên", "") or "").strip(),
                "unit": str(row.get("Đơn vị", "") or "").strip(),
                "day1": check if len(days) > 0 else "",
                "day2": check if len(days) > 1 else "",
                "day3": check if len(days) > 2 else "",
                "note": ""
            })
        students = [s for s in students if s["id"] or s["name"]]
        num_attended = sum(
            1 for s in students if "X" in [s.get("day1", ""), s.get("day2", ""), s.get("day3", "")]
        )
        with open("attendance_template.html", "r", encoding="utf-8") as f:
            template_str = f.read()
        template = Template(template_str)
        attendance_html = template.render(
            students=students,
            course_name=course_name,
            training_type=training_type,
            time=time,
            location=location,
            num_total=len(students),
            num_attended=num_attended,
            class_code=class_info.get("class_code", ""),
            gv_huong_dan=gv_huong_dan,
            days=days,
            logo_base64=logo_base64,
            max_rows_per_page=23
        )
        attendance_html_with_print = """
        <div style="text-align:right; margin-bottom:12px;" class="no-print">
            <button onclick="window.print()" style="font-size:18px;padding:6px 18px;">🖨️ In bảng điểm danh</button>
        </div>
        """ + attendance_html
        st.components.v1.html(attendance_html_with_print, height=1000, scrolling=True)

